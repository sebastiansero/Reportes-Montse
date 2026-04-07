from pathlib import Path

from openpyxl import Workbook

from src.config import ConfigLoader
from src.data_loader import ExcelLoader


def test_template_file_is_flagged_as_output_template(tmp_path):
    path = tmp_path / "PLANTILLA EMISIONES.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "EJECUTIVOS"
    ws.cell(1, 1, "FECHA EMISION")
    ws.cell(1, 2, "N° DE POLIZA")
    ws.cell(1, 3, "NOMBRE DE AGENTE")
    wb.save(path)

    config = ConfigLoader()
    loader = ExcelLoader(config)

    with path.open("rb") as file_obj:
        class UploadedFile:
            name = Path(file_obj.name).name

            def read(self):
                return file_obj.read()

            def seek(self, value):
                return file_obj.seek(value)

        df, errors = loader.load_and_validate([UploadedFile()], "emision_mensual")

    assert df.empty
    assert errors
    assert "plantilla de salida" in errors[0].lower()

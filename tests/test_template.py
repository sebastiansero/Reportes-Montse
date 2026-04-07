from io import BytesIO

import pandas as pd
from openpyxl import Workbook, load_workbook

from src.template_engine import TemplateFiller


def test_generic_template_fill(tmp_path):
    template_path = tmp_path / "template_operativo.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "COTIZACIONES"
    ws.cell(1, 1, "#")
    ws.cell(1, 2, "EJECUTIVO DE CUENTA")
    ws.cell(1, 3, "PRIMA TOTAL")
    wb.save(template_path)

    df = pd.DataFrame([{"EJECUTIVO DE CUENTA": "Hazel Castro", "PRIMA TOTAL": 8410.66}])
    report_config = {
        "template_path": str(template_path),
        "template_sheet": "COTIZACIONES",
        "template_header_row": 1,
        "template_data_row": 2,
        "source_type": "cotizaciones",
    }

    filler = TemplateFiller(str(tmp_path))
    output = filler.fill_template(df, report_config)
    loaded = load_workbook(BytesIO(output.getvalue()))
    ws = loaded["COTIZACIONES"]

    assert ws.cell(2, 1).value == 1
    assert ws.cell(2, 2).value == "Hazel Castro"
    assert ws.cell(2, 3).value == 8410.66


def test_renovaciones_template_preserves_formula_columns(tmp_path):
    template_path = tmp_path / "template_renovaciones.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "CONCENTRADO"
    headers = [
        None,
        "FECHA VTO",
        "HOY",
        "VENCIDA",
        "POLIZA",
        "COD ASEG",
        "NOMBRE DEL ASEG.",
        "COD. AGTE.",
        "NOMBRE DEL AGTE.",
        "MON.",
        "PRIMA TOTAL",
        "STROS",
        "RENOVACION",
        "EJECUTIVO",
        "EJECUTIVO",
        "RENOVADA",
        "ESTATUS",
        "COMENTARIOS",
    ]
    for idx, header in enumerate(headers, start=1):
        ws.cell(1, idx, header)
        ws.cell(2, idx, None)
    wb.save(template_path)

    df = pd.DataFrame(
        [
            {
                "FECHA VTO": pd.Timestamp("2026-04-28"),
                "POLIZA": "005233258",
                "COD ASEG": "002847464",
                "NOMBRE DEL ASEG.": "CARMEN RAMIREZ GARCI",
                "COD. AGTE.": "00355",
                "NOMBRE DEL AGTE.": "ESTEBAN FUENTES NIVON",
                "MON.": "$",
                "PRIMA TOTAL": 9354,
                "STROS": "",
                "RENOVACION": 0,
                "EJECUTIVO": "Hazel Castro",
                "RENOVADA": "",
                "COMENTARIOS": "",
            }
        ]
    )
    report_config = {
        "template_path": str(template_path),
        "template_sheet": "CONCENTRADO",
        "template_header_row": 1,
        "template_data_row": 2,
        "source_type": "renovaciones",
    }

    filler = TemplateFiller(str(tmp_path))
    output = filler.fill_template(df, report_config)
    loaded = load_workbook(BytesIO(output.getvalue()))
    ws = loaded["CONCENTRADO"]

    assert ws.cell(2, 2).value.year == 2026
    assert ws.cell(2, 15).value == "Hazel Castro"
    assert ws.cell(2, 3).value == "=TODAY()"
    assert ws.cell(2, 4).value == '=IF(B2="","",IF(B2<=C2,"SI","NO"))'


def test_emisiones_template_supports_custom_workbook_layout(tmp_path):
    template_path = tmp_path / "template_emisiones.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "DESGLOCE"
    for col in range(2, 11):
        ws.cell(4, col, None)
    ws_exec = wb.create_sheet("EJECUTIVOS")
    ws_direct = wb.create_sheet("CLAVE DIRECTA")

    headers = [
        "FECHA EMISION",
        "N° EMISION",
        "N° DE POLIZA",
        "TIPO DE NEGOCIO",
        "FECHA INICIO VIGENCIA",
        "FECHA FIN VIGENCIA",
        "COBERTURA",
        "ASEGURADO",
        "COD. AGENTE",
        "NOMBRE DE AGENTE",
        "PRIMA TOTAL",
        "FORMA PAGO",
        "PAGADO / NO PAGADO",
        "EJECUTIVO",
        "COMENTARIOS",
    ]

    for sheet in [ws_exec, ws_direct]:
        sheet.cell(1, 1, "MARZO 2026")
        for idx, header in enumerate(headers, start=1):
            sheet.cell(2, idx, header)
            sheet.cell(3, idx, None)

    df = pd.DataFrame(
        [
            {
                "EJECUTIVO DE CUENTA": "Hazel Castro",
                "FECHA EMISION": pd.Timestamp("2026-04-02"),
                "CLAVE DE AGENTE": "00355",
                "NOMBRE DE AGENTE": "ESTEBAN FUENTES NIVON",
                "ASEGURADO": "CARMEN RAMIREZ GARCI",
                "TIPO DE NEGOCIO": "INDIVIDUAL",
                "PRIMA TOTAL": 9354,
                "FORMA PAGO": "CONTADO",
                "FECHA INICIO VIGENCIA": pd.Timestamp("2026-04-02"),
                "FECHA FIN VIGENCIA": pd.Timestamp("2027-04-01"),
                "N° DE POLIZA": "005233258",
                "EMISION / RENOVACION / REEXPEDICION / CANCELACION": "EMISION",
                "COMENTARIOS": "Renueva a 005233999",
            },
            {
                "EJECUTIVO DE CUENTA": "Claudia Rios",
                "FECHA EMISION": pd.Timestamp("2026-04-03"),
                "CLAVE DE AGENTE": "20087",
                "NOMBRE DE AGENTE": "ANA ASESORIA Y PROTECCION PATRIMONIAL SC",
                "ASEGURADO": "ASEGURADO CLAVE DIRECTA",
                "TIPO DE NEGOCIO": "FLOTILLA",
                "PRIMA TOTAL": 15000,
                "FORMA PAGO": "MENSUAL",
                "FECHA INICIO VIGENCIA": pd.Timestamp("2026-04-03"),
                "FECHA FIN VIGENCIA": pd.Timestamp("2027-04-02"),
                "N° DE POLIZA": "005244000",
                "EMISION / RENOVACION / REEXPEDICION / CANCELACION": "EMISION",
                "COMENTARIOS": "",
            },
            {
                "EJECUTIVO DE CUENTA": "Claudia Rios",
                "FECHA EMISION": pd.Timestamp("2026-04-04"),
                "CLAVE DE AGENTE": "20087",
                "NOMBRE DE AGENTE": "ANA ASESORIA Y PROTECCION PATRIMONIAL SC",
                "ASEGURADO": "ASEGURADO RENOVACION",
                "TIPO DE NEGOCIO": "FLOTILLA",
                "PRIMA TOTAL": 15200,
                "FORMA PAGO": "MENSUAL",
                "FECHA INICIO VIGENCIA": pd.Timestamp("2026-04-04"),
                "FECHA FIN VIGENCIA": pd.Timestamp("2027-04-03"),
                "N° DE POLIZA": "005255000",
                "EMISION / RENOVACION / REEXPEDICION / CANCELACION": "RENOVACION",
                "COMENTARIOS": "Renueva a 005000001",
            },
        ]
    )

    report_config = {
        "template_path": str(template_path),
        "template_sheet": "EJECUTIVOS",
        "template_header_row": 2,
        "template_data_row": 3,
        "source_type": "emisiones",
    }

    wb.save(template_path)
    filler = TemplateFiller(str(tmp_path))
    output = filler.fill_template(df, report_config, report_key="emision_mensual")
    loaded = load_workbook(BytesIO(output.getvalue()))

    ws_exec = loaded["EJECUTIVOS"]
    ws_direct = loaded["CLAVE DIRECTA"]
    ws_summary = loaded["DESGLOCE"]

    assert ws_exec["A1"].value == "ABRIL 2026"
    assert ws_exec.cell(3, 3).value == "005233258"
    assert ws_exec.cell(3, 14).value == "Hazel Castro"
    assert ws_exec.cell(4, 3).value == "005255000"
    assert ws_exec.cell(4, 14).value == "Claudia Rios"
    assert ws_direct.cell(3, 3).value == "005244000"
    assert ws_direct.cell(3, 14).value == "CLAVE DIRECTA"
    assert ws_summary.cell(4, 8).value == "CLAVE DIRECTA"


def test_combined_workbook_contains_each_report_sheet(tmp_path):
    operativo_path = tmp_path / "template_operativo.xlsx"
    renovaciones_path = tmp_path / "template_renovaciones.xlsx"

    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "COTIZACIONES"
    ws1.cell(1, 1, "#")
    ws1.cell(1, 2, "EJECUTIVO DE CUENTA")
    wb1.create_sheet("EMISIONES Y CANCELACIONES")
    ws_em = wb1["EMISIONES Y CANCELACIONES"]
    ws_em.cell(2, 1, "#")
    ws_em.cell(2, 2, "EJECUTIVO DE CUENTA")
    wb1.save(operativo_path)

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "CONCENTRADO"
    ws2.cell(1, 2, "FECHA VTO")
    ws2.cell(1, 14, "EJECUTIVO")
    ws2.cell(1, 15, "EJECUTIVO")
    wb2.save(renovaciones_path)

    class TestConfig:
        def __init__(self):
            self._reports = {
                "cotizaciones": {
                    "template_path": str(operativo_path),
                    "template_sheet": "COTIZACIONES",
                    "template_header_row": 1,
                    "template_data_row": 2,
                    "combined_sheet_name": "Cotizaciones",
                    "source_type": "cotizaciones",
                    "name": "Cotizaciones",
                },
                "emision_mensual": {
                    "template_path": str(operativo_path),
                    "template_sheet": "EMISIONES Y CANCELACIONES",
                    "template_header_row": 2,
                    "template_data_row": 3,
                    "combined_sheet_name": "Emisiones",
                    "source_type": "emisiones",
                    "name": "Emisiones",
                },
                "renovaciones": {
                    "template_path": str(renovaciones_path),
                    "template_sheet": "CONCENTRADO",
                    "template_header_row": 1,
                    "template_data_row": 2,
                    "combined_sheet_name": "Renovaciones",
                    "source_type": "renovaciones",
                    "name": "Renovaciones",
                },
            }

        def get_report_config(self, report_key):
            return self._reports.get(report_key)

    datasets = {
        "cotizaciones": pd.DataFrame([{"EJECUTIVO DE CUENTA": "Hazel Castro"}]),
        "emision_mensual": pd.DataFrame([{"EJECUTIVO DE CUENTA": "Hazel Castro"}]),
        "renovaciones": pd.DataFrame([{"FECHA VTO": pd.Timestamp("2026-04-28"), "EJECUTIVO": "Hazel Castro"}]),
    }

    filler = TemplateFiller(str(tmp_path))
    output = filler.fill_combined_report(datasets, TestConfig())
    loaded = load_workbook(BytesIO(output.getvalue()))

    assert "Cotizaciones" in loaded.sheetnames
    assert "Emisiones" in loaded.sheetnames
    assert "Renovaciones" in loaded.sheetnames

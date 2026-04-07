from copy import copy
import io
import os

import pandas as pd
from loguru import logger
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from .agent_catalog import clean_text, normalize_text


SPANISH_MONTH_NAMES = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}


class TemplateFiller:
    def __init__(self, template_dir: str = "templates"):
        self.template_dir = template_dir
        os.makedirs(self.template_dir, exist_ok=True)

    def fill_template(
        self,
        df: pd.DataFrame,
        report_config: dict,
        report_key: str | None = None,
        template_bytes: bytes | None = None,
    ) -> io.BytesIO:
        template_path = report_config.get("template_path", "")
        if not template_path and template_bytes is None:
            raise ValueError("No template path configured for this report.")

        try:
            wb = self._load_template(template_path, template_bytes)
            template_mode = self._resolve_template_mode(wb, report_config, report_key)

            if template_mode == "emisiones":
                self._fill_emisiones_template(wb, df)
            elif template_mode == "renovaciones":
                self._fill_renovaciones_template(wb, df, report_config)
            else:
                self._fill_generic_template(wb, df, report_config)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            logger.info("Successfully filled template {}", template_path or "uploaded workbook")
            return output
        except Exception as exc:
            logger.error("Error filling template {}: {}", template_path or "uploaded workbook", exc)
            raise

    def fill_combined_report(self, datasets: dict[str, pd.DataFrame], config_loader) -> io.BytesIO:
        combined = Workbook()
        default_sheet = combined.active
        combined.remove(default_sheet)

        for report_key, df in datasets.items():
            if df is None or df.empty:
                continue

            report_config = config_loader.get_report_config(report_key)
            if not report_config:
                continue

            rendered = self.fill_template(df, report_config, report_key=report_key)
            source_wb = load_workbook(io.BytesIO(rendered.getvalue()))
            source_sheet_name = report_config.get("template_sheet")
            source_ws = source_wb[source_sheet_name] if source_sheet_name in source_wb.sheetnames else source_wb.active
            target_title = self._safe_sheet_title(report_config.get("combined_sheet_name") or report_config.get("name") or report_key)
            target_ws = combined.create_sheet(title=target_title)
            self._copy_sheet_contents(source_ws, target_ws)

        if not combined.sheetnames:
            combined.create_sheet(title="Reporte")

        output = io.BytesIO()
        combined.save(output)
        output.seek(0)
        return output

    def _load_template(self, template_path: str, template_bytes: bytes | None):
        if template_bytes is not None:
            return load_workbook(io.BytesIO(template_bytes))

        if not os.path.exists(template_path):
            logger.warning("Template {} not found. Generating default.", template_path)
            self._generate_default_template(pd.DataFrame(), template_path)

        return load_workbook(template_path)

    def _resolve_template_mode(self, wb, report_config: dict, report_key: str | None) -> str:
        sheet_names = set(wb.sheetnames)
        if report_key == "emision_mensual" and {"EJECUTIVOS", "CLAVE DIRECTA"}.issubset(sheet_names):
            return "emisiones"

        if report_config.get("source_type") == "renovaciones" and "CONCENTRADO" in sheet_names:
            return "renovaciones"

        return "generic"

    def _fill_generic_template(self, wb, df: pd.DataFrame, report_config: dict) -> None:
        sheet_name = report_config.get("template_sheet")
        header_row = report_config.get("template_header_row", 1)
        data_start_row = report_config.get("template_data_row", header_row + 1)

        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        header_map = self._build_header_map(ws, header_row)
        data_map = {normalize_text(column): column for column in df.columns}

        self._clear_rows(ws, data_start_row, len(df))

        for offset, (_, record) in enumerate(df.iterrows(), start=0):
            row_index = data_start_row + offset
            self._copy_row_style(ws, data_start_row, row_index)
            for col_index in range(1, ws.max_column + 1):
                header_key = header_map.get(col_index, "")
                if not header_key:
                    continue
                if header_key == "#":
                    ws.cell(row=row_index, column=col_index, value=offset + 1)
                    continue
                source_column = data_map.get(header_key)
                if not source_column:
                    continue
                ws.cell(row=row_index, column=col_index, value=self._excel_safe_value(record[source_column]))

    def _fill_emisiones_template(self, wb, df: pd.DataFrame) -> None:
        ejecutivo_df = df[df["EJECUTIVO DE CUENTA"].fillna("").astype(str).str.strip() != ""].copy()
        directa_df = df[df["EJECUTIVO DE CUENTA"].fillna("").astype(str).str.strip() == ""].copy()
        directa_df.loc[:, "EJECUTIVO DE CUENTA"] = "CLAVE DIRECTA"

        if "EJECUTIVOS" in wb.sheetnames:
            self._fill_emisiones_sheet(wb["EJECUTIVOS"], ejecutivo_df, month_style="full")
        if "CLAVE DIRECTA" in wb.sheetnames:
            self._fill_emisiones_sheet(wb["CLAVE DIRECTA"], directa_df, month_style="split")
        if "DESGLOCE" in wb.sheetnames:
            self._fill_emisiones_summary(wb["DESGLOCE"], df)

    def _fill_emisiones_sheet(self, ws, df: pd.DataFrame, month_style: str) -> None:
        header_row = 2
        data_start_row = 3
        header_positions = self._header_positions(ws, header_row)

        self._set_emisiones_title(ws, df, month_style)
        self._clear_rows(ws, data_start_row, len(df))

        for offset, (_, record) in enumerate(df.iterrows(), start=0):
            row_index = data_start_row + offset
            self._copy_row_style(ws, data_start_row, row_index)
            values = {
                "FECHA EMISION": record.get("FECHA EMISION"),
                "N° EMISION": "",
                "N° DE POLIZA": record.get("N° DE POLIZA"),
                "TIPO DE NEGOCIO": record.get("TIPO DE NEGOCIO"),
                "FECHA INICIO VIGENCIA": record.get("FECHA INICIO VIGENCIA"),
                "FECHA FIN VIGENCIA": record.get("FECHA FIN VIGENCIA"),
                "COBERTURA": "",
                "ASEGURADO": record.get("ASEGURADO"),
                "COD. AGENTE": record.get("CLAVE DE AGENTE"),
                "NOMBRE DE AGENTE": record.get("NOMBRE DE AGENTE"),
                "PRIMA TOTAL": record.get("PRIMA TOTAL"),
                "FORMA PAGO": record.get("FORMA PAGO"),
                "PAGADO / NO PAGADO": "",
                "EJECUTIVO": record.get("EJECUTIVO DE CUENTA"),
                "COMENTARIOS": record.get("COMENTARIOS"),
            }

            for header, value in values.items():
                col_index = self._first_column(header_positions, header)
                if col_index:
                    ws.cell(row=row_index, column=col_index, value=self._excel_safe_value(value))

    def _set_emisiones_title(self, ws, df: pd.DataFrame, month_style: str) -> None:
        if "FECHA EMISION" not in df.columns:
            return

        fechas = pd.to_datetime(df["FECHA EMISION"], errors="coerce").dropna()
        if fechas.empty:
            return

        ref_date = fechas.iloc[0]
        month_name = SPANISH_MONTH_NAMES.get(ref_date.month, "")
        if not month_name:
            return

        if month_style == "full":
            ws["A1"] = f"{month_name} {ref_date.year}"
        else:
            ws["A1"] = month_name
            ws["B1"] = ref_date.year

    def _fill_emisiones_summary(self, ws, df: pd.DataFrame) -> None:
        detail = df.copy()
        detail["RESUMEN_EJECUTIVO"] = detail["EJECUTIVO DE CUENTA"].fillna("").map(clean_text).replace("", "CLAVE DIRECTA")
        detail["RESUMEN_AGENTE"] = detail["NOMBRE DE AGENTE"].fillna("").map(clean_text).replace("", "SIN AGENTE")

        agent_summary = (
            detail.groupby(["RESUMEN_EJECUTIVO", "RESUMEN_AGENTE"], dropna=False)
            .size()
            .reset_index(name="TOTAL")
            .sort_values(["RESUMEN_EJECUTIVO", "RESUMEN_AGENTE"])
        )
        executive_summary = (
            detail.groupby(["RESUMEN_EJECUTIVO"], dropna=False)
            .size()
            .reset_index(name="EMISIONES")
            .sort_values(["RESUMEN_EJECUTIVO"])
        )

        last_needed_row = max(4 + len(agent_summary), 4 + len(executive_summary) + 1, ws.max_row)
        for row_index in range(3, last_needed_row + 1):
            self._copy_row_style(ws, 4, row_index)
            for col_index in range(2, 11):
                ws.cell(row=row_index, column=col_index).value = None

        ws.cell(3, 2, "EJECUTIVO")
        ws.cell(3, 3, "NOMBRE DE AGENTE")
        ws.cell(3, 4, "REGISTROS")
        ws.cell(3, 8, "EJECUTIVO")
        ws.cell(3, 9, "EMISIONES")
        ws.cell(3, 10, "TOTAL")

        for offset, (_, row) in enumerate(agent_summary.iterrows(), start=0):
            row_index = 4 + offset
            ws.cell(row_index, 2, row["RESUMEN_EJECUTIVO"])
            ws.cell(row_index, 3, row["RESUMEN_AGENTE"])
            ws.cell(row_index, 4, int(row["TOTAL"]))

        running_total = 0
        for offset, (_, row) in enumerate(executive_summary.iterrows(), start=0):
            row_index = 4 + offset
            running_total += int(row["EMISIONES"])
            ws.cell(row_index, 8, row["RESUMEN_EJECUTIVO"])
            ws.cell(row_index, 9, int(row["EMISIONES"]))
            ws.cell(row_index, 10, running_total)

        total_row = 4 + len(executive_summary)
        ws.cell(total_row, 8, "Total general")
        ws.cell(total_row, 9, int(executive_summary["EMISIONES"].sum()) if not executive_summary.empty else 0)
        ws.cell(total_row, 10, int(executive_summary["EMISIONES"].sum()) if not executive_summary.empty else 0)

    def _fill_renovaciones_template(self, wb, df: pd.DataFrame, report_config: dict) -> None:
        sheet_name = report_config.get("template_sheet", "CONCENTRADO")
        header_row = report_config.get("template_header_row", 1)
        data_start_row = report_config.get("template_data_row", 2)

        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        header_positions = self._header_positions(ws, header_row)
        get_position = lambda header: header_positions.get(normalize_text(header), [None])[0]

        target_last_row = max(ws.max_row, data_start_row + len(df) - 1)
        for row_index in range(data_start_row, target_last_row + 1):
            self._copy_row_style(ws, data_start_row, row_index)
            for col_index in range(2, 19):
                ws.cell(row=row_index, column=col_index).value = None

        ejecutivo_columns = header_positions.get(normalize_text("EJECUTIVO"), [])
        ejecutivo_manual_col = ejecutivo_columns[1] if len(ejecutivo_columns) > 1 else (ejecutivo_columns[0] if ejecutivo_columns else None)

        column_map = {
            "FECHA VTO": get_position("FECHA VTO"),
            "POLIZA": get_position("POLIZA"),
            "COD ASEG": get_position("COD ASEG"),
            "NOMBRE DEL ASEG.": get_position("NOMBRE DEL ASEG."),
            "COD. AGTE.": get_position("COD. AGTE."),
            "NOMBRE DEL AGTE.": get_position("NOMBRE DEL AGTE."),
            "MON.": get_position("MON."),
            "PRIMA TOTAL": get_position("PRIMA TOTAL"),
            "STROS": get_position("STROS"),
            "RENOVACION": get_position("RENOVACION"),
            "EJECUTIVO": ejecutivo_manual_col,
            "RENOVADA": get_position("RENOVADA"),
            "COMENTARIOS": get_position("COMENTARIOS"),
        }

        for offset, (_, record) in enumerate(df.iterrows(), start=0):
            row_index = data_start_row + offset
            self._copy_row_style(ws, data_start_row, row_index)

            for column_name, col_index in column_map.items():
                if not col_index or column_name not in record.index:
                    continue
                ws.cell(row=row_index, column=col_index, value=self._excel_safe_value(record[column_name]))

            ws.cell(row=row_index, column=3, value="=TODAY()")
            ws.cell(row=row_index, column=4, value=f'=IF(B{row_index}="","",IF(B{row_index}<=C{row_index},"SI","NO"))')
            ws.cell(row=row_index, column=14, value=f'=IF(P{row_index}="SI","",O{row_index})')
            ws.cell(row=row_index, column=17, value=f'=IF(P{row_index}="SI","RENOVADA","PENDIENTE")')

    def _build_header_map(self, ws, header_row: int) -> dict[int, str]:
        header_map = {}
        for col_index in range(1, ws.max_column + 1):
            raw_value = ws.cell(row=header_row, column=col_index).value
            raw_text = str(raw_value).replace("\n", " ") if raw_value is not None else ""
            cleaned = "#" if raw_text.strip() == "#" else normalize_text(raw_text)
            header_map[col_index] = cleaned
        return header_map

    def _header_positions(self, ws, header_row: int) -> dict[str, list[int]]:
        positions: dict[str, list[int]] = {}
        for col_index in range(1, ws.max_column + 1):
            raw_value = ws.cell(row=header_row, column=col_index).value
            raw_text = str(raw_value).replace("\n", " ") if raw_value is not None else ""
            cleaned = "#" if raw_text.strip() == "#" else normalize_text(raw_text)
            if not cleaned:
                continue
            positions.setdefault(cleaned, []).append(col_index)
        return positions

    def _first_column(self, header_positions: dict[str, list[int]], header_name: str) -> int | None:
        positions = header_positions.get(normalize_text(header_name), [])
        return positions[0] if positions else None

    def _clear_rows(self, ws, data_start_row: int, incoming_rows: int) -> None:
        last_row = max(ws.max_row, data_start_row + incoming_rows - 1)
        for row_index in range(data_start_row, last_row + 1):
            self._copy_row_style(ws, data_start_row, row_index)
            for col_index in range(1, ws.max_column + 1):
                ws.cell(row=row_index, column=col_index).value = None

    def _copy_row_style(self, ws, template_row: int, target_row: int) -> None:
        if template_row == target_row or template_row > ws.max_row:
            return
        for col_index in range(1, ws.max_column + 1):
            source = ws.cell(row=template_row, column=col_index)
            target = ws.cell(row=target_row, column=col_index)
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)

    def _copy_sheet_contents(self, source_ws, target_ws) -> None:
        for col_letter, dimension in source_ws.column_dimensions.items():
            target_ws.column_dimensions[col_letter].width = dimension.width
            target_ws.column_dimensions[col_letter].hidden = dimension.hidden

        for row_idx, dimension in source_ws.row_dimensions.items():
            target_ws.row_dimensions[row_idx].height = dimension.height
            target_ws.row_dimensions[row_idx].hidden = dimension.hidden

        for row in source_ws.iter_rows():
            for source_cell in row:
                if source_cell.__class__.__name__ == "MergedCell":
                    continue
                target_cell = target_ws.cell(row=source_cell.row, column=source_cell.column, value=source_cell.value)
                if source_cell.has_style:
                    target_cell.font = copy(source_cell.font)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.border = copy(source_cell.border)
                    target_cell.alignment = copy(source_cell.alignment)
                    target_cell.number_format = source_cell.number_format
                    target_cell.protection = copy(source_cell.protection)

        for merged_range in source_ws.merged_cells.ranges:
            target_ws.merge_cells(str(merged_range))

    def _safe_sheet_title(self, title: str) -> str:
        clean = "".join(ch for ch in title if ch not in '[]:*?/\\')
        return (clean or "Reporte")[:31]

    def _excel_safe_value(self, value):
        if pd.isna(value):
            return None
        if isinstance(value, pd.Timestamp):
            return value.to_pydatetime()
        return value

    def _generate_default_template(self, df: pd.DataFrame, path: str):
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"

        headers = df.columns.tolist()
        if headers:
            ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        wb.save(path)
